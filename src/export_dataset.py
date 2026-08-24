from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.bayut_api import BayutAlgoliaClient
from src.config import (
    ALGOLIA_APP_ID,
    ALGOLIA_SALE_INDEX,
    ALGOLIA_SEARCH_API_KEY,
    DATASET_CSV,
    DATASET_JSONL,
    DATASET_XLSX,
    OUTPUT_DIR,
    RAW_DIR,
    ensure_directories,
)
from src.detail_parser import BayutDetailParser
from src.models import HousingListing
from src.parser import BayutParser


DETAILS_DIR = RAW_DIR / "details"

ORDERED_COLUMNS = [
    # Identity & Provenance
    "listing_id",
    "url",
    "source_discovery",
    "detail_html_captured",
    "description_source",
    # Group A
    "purpose",
    "property_type",
    "price",
    "price_period",
    "currency",
    "bedrooms",
    "bathrooms",
    "area_sqm",
    "location_raw",
    "agency_name",
    "is_verified",
    "date_listed",
    "description_raw",
    "language",
    # Group B
    "compound_name",
    "developer_name",
    "governorate",
    "city",
    "district",
    "finishing_level",
    "delivery_status",
    "delivery_date",
    "sale_type",
    "payment_type",
    "down_payment_amount",
    "down_payment_pct",
    "installment_years",
    "installment_amount",
    "installment_frequency",
    "cash_discount_pct",
    "amenities",
    "floor_number",
    "garden_area_sqm",
    "roof_area_sqm",
    "is_negotiable",
    # Derived
    "price_per_sqm",
    "total_installment_cost",
]


def build_dataset(
    sale_count: int = 280,
    rent_count: int = 280,
) -> list[HousingListing]:
    ensure_directories()
    DETAILS_DIR.mkdir(parents=True, exist_ok=True)

    client = BayutAlgoliaClient(
        ALGOLIA_APP_ID,
        ALGOLIA_SEARCH_API_KEY,
        ALGOLIA_SALE_INDEX,
    )
    detail_parser = BayutDetailParser()
    parser = BayutParser()

    raw_listings: list[dict[str, Any]] = []

    try:
        # Fetch sale hits
        pages_needed_sale = (sale_count + 99) // 100
        for page in range(pages_needed_sale):
            hits = client.get_sale_hits(page=page, hits_per_page=100)
            raw_listings.extend(hits[: max(0, sale_count - len([x for x in raw_listings if x.get("purpose") == "for-sale"]))])

        # Fetch rent hits
        pages_needed_rent = (rent_count + 99) // 100
        for page in range(pages_needed_rent):
            hits = client.get_rent_hits(page=page, hits_per_page=100)
            raw_listings.extend(hits[: max(0, (sale_count + rent_count) - len(raw_listings))])

    finally:
        client.close()

    # Deduplicate by objectID or externalID
    seen_ids = set()
    unique_raw: list[dict[str, Any]] = []
    for hit in raw_listings:
        hit_id = str(hit.get("objectID") or hit.get("externalID") or hit.get("id"))
        if hit_id not in seen_ids:
            seen_ids.add(hit_id)
            unique_raw.append(hit)

    total = len(unique_raw)
    print(f"Total Unique Candidate Listings: {total}")

    listings: list[HousingListing] = []

    for idx, raw in enumerate(unique_raw, 1):
        if idx % 10 == 0 or idx == total:
            print(f"  ▶ Parsing listings: [{idx}/{total}] ({(idx/total)*100:.1f}%)...", end="\r", flush=True)

        html_file = None
        for candidate_id in (
            str(raw.get("objectID") or ""),
            str(raw.get("id") or ""),
            str(raw.get("externalID") or ""),
            str(raw.get("listing_id") or ""),
        ):
            if candidate_id and candidate_id != "None":
                f = DETAILS_DIR / f"{candidate_id}.html"
                if f.exists() and f.stat().st_size > 5000:
                    html_file = f
                    break

        detail_data: dict[str, Any] | None = None
        if html_file and html_file.exists():
            try:
                detail_data = detail_parser.parse_file(str(html_file))
            except Exception as e:
                print(f"\n[WARN] Error parsing {html_file.name}: {e}")

        listing = parser.parse(raw, detail_data)
        listings.append(listing)

    print()
    return listings


def export_to_excel(df: pd.DataFrame, output_path: Path) -> None:
    """
    Export DataFrame to beautifully styled Excel spreadsheet.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Bayut Egyptian Housing", index=False)
        worksheet = writer.sheets["Bayut Egyptian Housing"]

        # Styles
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        border_thin = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Format header
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        worksheet.row_dimensions[1].height = 28

        # Format data rows
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            is_even = (row_num % 2 == 0)
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = border_thin
                if is_even:
                    cell.fill = zebra_fill

                # Number formatting
                col_name = df.columns[col_num - 1]
                if col_name in ("price", "down_payment_amount", "installment_amount", "price_per_sqm", "total_installment_cost") and cell.value is not None:
                    try:
                        cell.number_format = "#,##0"
                    except Exception:
                        pass
                elif col_name in ("down_payment_pct", "cash_discount_pct") and cell.value is not None:
                    try:
                        cell.number_format = "0.0%"
                    except Exception:
                        pass
                elif col_name in ("area_sqm", "garden_area_sqm", "roof_area_sqm", "installment_years") and cell.value is not None:
                    try:
                        cell.number_format = "#,##0.0"
                    except Exception:
                        pass

        # Auto-fit column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = min(len(val), 50)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)


def main() -> None:
    print("Building canonical Egyptian housing dataset from Algolia + Detail HTML...")
    listings = build_dataset(sale_count=280, rent_count=280)

    # Convert to dictionaries
    records = []
    for item in listings:
        d = item.model_dump()
        # Convert list of amenities to comma-separated string for Excel/CSV
        if isinstance(d.get("amenities"), list):
            d["amenities"] = ", ".join(d["amenities"])
        records.append(d)

    df = pd.DataFrame(records)

    # Reorder columns
    present_cols = [c for c in ORDERED_COLUMNS if c in df.columns]
    df = df[present_cols]

    # Export formats
    print(f"\nExporting {len(df)} records to {DATASET_XLSX}...")
    export_to_excel(df, DATASET_XLSX)

    print(f"Exporting CSV to {DATASET_CSV}...")
    df.to_csv(DATASET_CSV, index=False, encoding="utf-8-sig")

    print(f"Exporting JSONL to {DATASET_JSONL}...")
    with open(DATASET_JSONL, "w", encoding="utf-8") as f:
        for item in listings:
            f.write(json.dumps(item.model_dump(), ensure_ascii=False) + "\n")

    # Export failure log
    from src.db import Database
    db = Database(OUTPUT_DIR.parent / "housing_pipeline.db")
    db.export_failures_csv(OUTPUT_DIR / "failure_log.csv")
    print(f"Exported failure log to {OUTPUT_DIR / 'failure_log.csv'}")

    print("\n" + "=" * 70)
    print("DATASET EXPORT COMPLETE")
    print("=" * 70)
    print(f"Total listings: {len(df)}")
    print(f"Sale listings: {sum(df['purpose'] == 'sale')}")
    print(f"Rent listings: {sum(df['purpose'] == 'rent')}")
    print(f"\nGovernorate Breakdown:\n{df['governorate'].value_counts(dropna=False).to_string()}")
    print(f"\nProperty Type Breakdown:\n{df['property_type'].value_counts(dropna=False).to_string()}")

    print("\nGroup B Field Coverage:")
    print("-" * 50)
    group_b = [
        "compound_name",
        "developer_name",
        "governorate",
        "city",
        "district",
        "finishing_level",
        "delivery_status",
        "delivery_date",
        "sale_type",
        "payment_type",
        "down_payment_amount",
        "down_payment_pct",
        "installment_years",
        "installment_amount",
        "installment_frequency",
        "cash_discount_pct",
        "amenities",
        "floor_number",
        "garden_area_sqm",
        "roof_area_sqm",
        "is_negotiable",
    ]
    for col in group_b:
        if col in df.columns:
            non_null = df[col].notna() & (df[col] != "")
            count = non_null.sum()
            pct = count / len(df) * 100
            print(f"{col:28} {count:>4}/{len(df):<4} ({pct:6.2f}%)")


if __name__ == "__main__":
    main()
